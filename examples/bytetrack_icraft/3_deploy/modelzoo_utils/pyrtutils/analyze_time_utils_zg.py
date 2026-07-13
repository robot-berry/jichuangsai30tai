import os
import pandas as pd
import numpy as np
import shutil
from matplotlib import pyplot as plt
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

def sheet_style(writer, max_index, df):
    workbook = writer.book
    worksheet = workbook['Sheet1']
    # 定义无边框样式
    border_style = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None)
    )

    # 去掉所有单元格的边框
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border_style

    ## 自动调整行宽
    for column in worksheet.columns:
        max_length = 4
        column_letter = get_column_letter(column[0].column)
        if column_letter == "B" : 
            for cell in column[:8]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        elif column_letter == "C":
            max_length = 15
            for cell in column[:8]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        else:
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 1) * 1.3
        worksheet.column_dimensions[column_letter].width = adjusted_width

    ## 居中对齐单元格内容
    alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment


    # 设置填充区域
    fill_style1 = PatternFill(start_color="C4DFA3", end_color="C4DFA3", fill_type="solid")
    cell_range1 = worksheet['A1:C50']
    fill_style2 = PatternFill(start_color="D1EAF0", end_color="D1EAF0", fill_type="solid")
    cell_range2 = worksheet['E1:J50']
    if max_index<30:
        max_index = 30
    else:
        max_index +=4 
    
    # 动态计算“内部数据”区域的范围
    internal_data_start_col_num = 13  # Column M
    # +1 because to_excel writes the index
    num_internal_data_cols = len(df.columns) + 1
    internal_data_end_col_num = internal_data_start_col_num + num_internal_data_cols - 1
    internal_data_end_col_letter = get_column_letter(internal_data_end_col_num)
    
    fill_style3 = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    cell_range3 = worksheet[f'M1:{internal_data_end_col_letter}{max_index}']

    # 设置填充颜色
    for row in cell_range1:
        for cell in row:
            cell.fill = fill_style1
    for row in cell_range2:
        for cell in row:
            cell.fill = fill_style2
    for row in cell_range3:
        for cell in row:
            cell.fill = fill_style3

    writer.save()
    writer.close()

def detect_file_type(filepath):
    _, ext = os.path.splitext(filepath)
    if ext.lower() == '.xlsx' or ext.lower() == '.xls':
        return 'excel'
    elif ext.lower() == '.txt':
        return 'txt'
    else:
        return 'unknown'
    
## 绘制各类hardop算子耗时堆叠柱状图
def draw_total_time(op_name,df_sum,filename,respath,resname):
    fig = plt.figure(figsize=(12, 8))
    plt.grid(ls="--", alpha=0.5)
    hard_time = list(df_sum["hard_time"])
    other_time = list(df_sum["other_time"])

    # 绘制堆叠柱状图
    plt.bar(op_name,hard_time,width=0.5,color="blue",label="hard_time")
    plt.bar(op_name,other_time,width=0.5,color="orange",label="other_time", bottom=hard_time)
    plt.xlabel('op_name')
    plt.ylabel('time(ms)')
    plt.legend()

    for i in range(len(op_name)):
        max_y = round(hard_time[i]+other_time[i],2)
        plt.text(op_name[i], max_y+0.2, max_y, va="bottom", ha="center")

    # 保存图片
    savepath = respath+filename+resname+'.png'
    plt.savefig(savepath, bbox_inches='tight')
    # plt.show()
    print("各类hardop算子耗时堆叠柱状图保存在", savepath)

## 绘制所有hardop算子耗时堆叠柱状图
def draw_per_op_time(op_id,df,filename,respath,resname):
    fig = plt.figure(figsize=(18, 10))
    plt.grid(ls="--", alpha=0.5)
    index = np.arange(len(op_id))
    width=0.5
    hard_time = list(df["hard_time"])
    other_time = list(df["other_time"])

    plt.bar(index,hard_time,width=width,color="blue",label="hard_time")
    plt.bar(index,other_time,width=width,color="orange",label="other_time",bottom=hard_time)#绘制柱状图

    plt.xticks(index, labels=op_id)
    plt.xticks(rotation= 45,fontsize=8)
    plt.xlabel('op_id')
    plt.ylabel('time(ms)')
    plt.legend()

    # 保存图片
    savepath = respath+filename+resname+'.png'
    plt.savefig(savepath, bbox_inches='tight')
    # plt.show()
    print("所有hardop算子的耗时堆叠柱状图保存在",savepath)

## 统计所有op的总耗时情况
def calculate_totaltime(writer,df,sheet_name,startcol=0,startrow=0):
    total_totaltime =  df["total_time"].sum()
    total_memcpytime =  df["memcpy_time"].sum()
    total_hardtime = df["hard_time"].sum()
    total_restime = df["other_time"].sum()

    df_total = df.copy()
    if 'is_io_process' in df_total.columns:
        df_total.loc['total']=["","","",total_totaltime,total_memcpytime,total_hardtime,total_restime, ""]
    else:
        df_total.loc['total']=["","","",total_totaltime,total_memcpytime,total_hardtime,total_restime]
    df_total = df_total.round(4)

    # 写入excel
    title_row = pd.DataFrame(index=['内部数据：'])
    title_row.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow)
    df_total.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow+2)
    writer.save()
    

## 统计所有HardOP算子细则：按照算子类别分类、统计时间
# def calculate_hardop_class_time(writer,df,sheet_name,startcol=12,startrow=0):
#     df_class = df.groupby(["op_name"]).sum().iloc[:,1:]
#     category_counts = df.groupby(["op_name"]).size().rename("op_nums")
#     df_class.insert(0, "op_nums", category_counts)
#     hardop_opnums = df_class["op_nums"].sum()
#     hardop_totaltime = df_class["total_time"].sum()
#     hardop_memcpytime = df_class["memcpy_time"].sum()
#     hardop_hardtime = df_class["hard_time"].sum()
#     hardop_restime = df_class["other_time"].sum()
    
#     # 检查 is_io_process 列是否存在，如果存在，则重命名为 io_op_num
#     if 'is_io_process' in df_class.columns:
#         df_class.rename(columns={'is_io_process': 'io_op_num'}, inplace=True)

#     # 剔除soft中memcpy_time
#     df_class["total_time"] = df_class["total_time"] - df_class["memcpy_time"]
#     hardop_totaltime = hardop_totaltime - hardop_memcpytime
#     df_class.drop(["memcpy_time"],axis=1,inplace=True)

#     # 检查 io_op_num 列是否存在，并相应地调整
#     if 'io_op_num' in df_class.columns:
#         io_process_totaltime = df_class["io_op_num"].sum()
#         df_class.loc['total_hardop']=[hardop_opnums,hardop_totaltime,hardop_hardtime,hardop_restime, io_process_totaltime]
#     else:
#         df_class.loc['total_hardop']=[hardop_opnums,hardop_totaltime,hardop_hardtime,hardop_restime]
        
#     df_class = df_class.round(4)
#     title_row = pd.DataFrame(index=['各类HardOP算子的耗时细则：'])
#     title_row.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow)
#     df_class.to_excel(writer, sheet_name=sheet_name, startcol=startcol,startrow=startrow+2)
#     writer.save()

## 统计所有HardOP算子细则：按照算子类别分类、统计时间
def calculate_hardop_class_time(writer,df,sheet_name,startcol=12,startrow=0):
    # 1. 兼容性检查：如果不存在 is_io_process 列，则执行旧的、兼容的逻辑
    if 'is_io_process' not in df.columns:
        df_class = df.groupby(["op_name"]).sum().iloc[:,1:]
        category_counts = df.groupby(["op_name"]).size().rename("op_nums")
        df_class.insert(0, "op_nums", category_counts)
        
        hardop_opnums = df_class["op_nums"].sum()
        hardop_totaltime = df_class["total_time"].sum()
        hardop_memcpytime = df_class["memcpy_time"].sum()
        hardop_hardtime = df_class["hard_time"].sum()
        hardop_restime = df_class["other_time"].sum()
        
        df_class["total_time"] = df_class["total_time"] - df_class["memcpy_time"]
        hardop_totaltime = hardop_totaltime - hardop_memcpytime
        df_class.drop(["memcpy_time"],axis=1,inplace=True)
        
        df_class.loc['total_hardop']=[hardop_opnums,hardop_totaltime,hardop_hardtime,hardop_restime]
        
        
        # 写入Excel - 使用旧的标题
        df_class = df_class.round(4)
        title_row = pd.DataFrame(index=['各类HardOP算子的耗时细则：'])
        title_row.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow)
        df_class.to_excel(writer, sheet_name=sheet_name, startcol=startcol,startrow=startrow+2)
        writer.save()
        
    # 2. 如果存在 is_io_process 列，则执行新的拆分逻辑
    else:
        # 2.1. 将数据拆分为 "主干算子" 和 "IO算子"
        df_backbone = df[df['is_io_process'] == False]
        df_io = df[df['is_io_process'] == True]

        # 2.2. 主干算子统计
        backbone_counts = df_backbone.groupby(["op_name"]).size().rename("backbone_op_num")
        backbone_times = df_backbone.groupby(["op_name"])[['total_time', 'hard_time', 'other_time', 'memcpy_time']].sum()
        
        # 合并主干算子统计结果
        df_backbone_class = pd.concat([backbone_counts, backbone_times], axis=1)
        
        # 剔除 memcpy_time
        df_backbone_class["total_time"] = df_backbone_class["total_time"] - df_backbone_class["memcpy_time"]
        df_backbone_class.drop(["memcpy_time"], axis=1, inplace=True)

        # 计算并添加 'total_backbone' 总结行
        total_backbone_row = df_backbone_class.sum()
        total_backbone_row.name = 'total_backbone'
        df_backbone_class = df_backbone_class.append(total_backbone_row)
        df_backbone_class['backbone_op_num'] = df_backbone_class['backbone_op_num'].astype(int)

        # 调整列顺序
        df_backbone_class = df_backbone_class[['backbone_op_num', 'total_time', 'hard_time', 'other_time']]
        df_backbone_class = df_backbone_class.round(4)

        # 写入主干算子表格
        title_row_backbone = pd.DataFrame(index=['主干HardOP算子的耗时细则：'])
        title_row_backbone.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow)
        df_backbone_class.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow+2)
        
        # 2.3. IO算子统计（如果存在IO算子）
        if not df_io.empty:
            io_counts = df_io.groupby(["op_name"]).size().rename("io_op_num")
            io_times = df_io.groupby(["op_name"])[['total_time', 'hard_time', 'other_time', 'memcpy_time']].sum()
            
            # 合并IO算子统计结果
            df_io_class = pd.concat([io_counts, io_times], axis=1)
            
            # 剔除 memcpy_time
            df_io_class["total_time"] = df_io_class["total_time"] - df_io_class["memcpy_time"]
            df_io_class.drop(["memcpy_time"], axis=1, inplace=True)

            # 计算并添加 'total_io' 总结行
            total_io_row = df_io_class.sum()
            total_io_row.name = 'total_io'
            df_io_class = df_io_class.append(total_io_row)
            df_io_class['io_op_num'] = df_io_class['io_op_num'].astype(int)

            # 调整列顺序
            df_io_class = df_io_class[['io_op_num', 'total_time', 'hard_time', 'other_time']]
            df_io_class = df_io_class.round(4)

            # 计算IO算子表格的起始行（在主干表格下方，留出适当间隔）
            io_startrow = startrow + len(df_backbone_class) + 5
            
            # 写入IO算子表格
            title_row_io = pd.DataFrame(index=['io算子的耗时细则：'])
            title_row_io.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=io_startrow)
            df_io_class.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=io_startrow+2)
        
        writer.save()


## 统计各类硬算子的时间，如imk、hardop、detpost、segpost
def calculate_hardop_time(writer,df,sheet_name,startcol=12,startrow=15):
    if 'is_io_process' in df.columns:
        df_main = df[df['is_io_process'] == False]
    else:
        df_main = df

    df_types = df_main.groupby(["op_type"]).sum()
    op_type = df_types.index
    columns_to_keep =[col for col in op_type if col.split("::")[0] == "customop" or col.split("::")[-1] == "HardOp"]
    df_hardop = df_types.loc[columns_to_keep,["total_time","memcpy_time","hard_time","other_time"]]
    total_hard_totaltime = df_hardop["total_time"].sum()
    total_hard_memcpytime = df_hardop["memcpy_time"].sum()
    total_hard_hardtime = df_hardop["hard_time"].sum()
    total_hard_restime = df_hardop["other_time"].sum()
    # 剔除soft中memcpy_time
    df_hardop["total_time"] = df_hardop["total_time"] - df_hardop["memcpy_time"]
    total_hard_totaltime = total_hard_totaltime - total_hard_memcpytime
    df_hardop.drop(["memcpy_time"],axis=1,inplace=True)

    df_hardop.loc['total'] = [total_hard_totaltime,total_hard_hardtime,total_hard_restime]
    df_hardop = df_hardop.round(4)

    # 写入excel
    title_row = pd.DataFrame(index=['各类硬算子耗时统计结果:'])
    title_row.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow)
    df_hardop.to_excel(writer, sheet_name=sheet_name, startcol=startcol,startrow=startrow+2)
    writer.save()

## 统计数据传入icore时间、icore耗时、传出icore时间、纯CPU端算子耗时
def calculate_detailtime(writer,df,sheet_name,startcol=12,startrow=25):
    cast_thresh = 0.001
    op_type = df.groupby(["op_type"]).sum().index
    # 从网络中获取IMK,POST
    IMK,POST,PRE,CENTER_CUSTOMOP = False,False,False,False
    pre_customop_time, other_customop_time= 0, 0
    imk_list,post_list,pre_list,other_customop_list = [],[],[],[]
    for col in op_type:
        a,b = col.split("::")[0],col.split("::")[-1]
        if a == "customop":
            if b == "ImageMake":
                IMK = True
                imk_list.append(col)
                print("network with IMK:",col)
            elif b[-4:] == "Post":
                POST = True
                post_list.append(col)
                print("network with POST:",col)
            elif "Pre" in b:
                PRE = True
                pre_list.append(col)
                print("network with PRE:",col)
                pre_customop_time += df.loc[df["op_type"]==col, "total_time"].values[0]
            else:
                CENTER_CUSTOMOP = True
                other_customop_list.append(col)
                other_customop_time += df.loc[df["op_type"]==col, "total_time"].values[0]
        else:
                continue
        
    # 若imk和post list为空，则为cdma搬数
    if imk_list == []:
        imk_list.append("cdma")
    if post_list == []:
        post_list.append("cdma")


    # 获取输入icore的时间
    if IMK:
        icore_in_time = df.loc[df["op_type"] == "customop::ImageMake", "hard_time"].values[0]
    else:
        icore_in_time = df.loc[df["op_type"] == "icraft::xir::HardOp", "memcpy_time"].sum()
       
    # 获取icore输出的时间   
    if POST:
        icore_out_time = df.loc[df["op_type"].str.split("::").str[-1].str[-4:] == "Post", "total_time"].values[0]
    else:
        filtered_cast_df = df[(df["op_type"] == "icraft::xir::Cast") & (df["memcpy_time"] > cast_thresh)]
        icore_out_time = filtered_cast_df["memcpy_time"].sum()
    
    # 获取icore的时间 
    # 如果存在 is_io_process 列，只统计主干算子（is_io_process=False）的时间
    
    if 'is_io_process' in df.columns:
        # 主干算子：只统计 is_io_process=False 的 HardOp
        df_backbone_hardop = df[(df["op_type"] == "icraft::xir::HardOp") & (df['is_io_process'] == False)]
       
        if IMK:
            icore_time = df_backbone_hardop["total_time"].sum()
        else:
            # 主干算子的memcpy_time（用于传入时间）
            backbone_memcpy = df_backbone_hardop["memcpy_time"].sum()
            icore_time = df_backbone_hardop["total_time"].sum() - backbone_memcpy
            # 更新 icore_in_time 为主干算子的 memcpy 时间
            if not IMK:
                icore_in_time = backbone_memcpy
    else:
        # 兼容旧逻辑：没有 is_io_process 列时，统计所有 HardOp
        if IMK:
            icore_time = df.loc[df["op_type"] == "icraft::xir::HardOp", "total_time"].sum()
        else:
            icore_time = df.loc[df["op_type"] == "icraft::xir::HardOp", "total_time"].sum() - icore_in_time

    # 若网络中间含有其它customop,需加上该硬算子时间    
    if CENTER_CUSTOMOP:
            icore_time = icore_time + other_customop_time
    
    # 纯CPU端算子耗时
    df_cpu = df[(df["op_type"] != "icraft::xir::HardOp") & (~df["op_type"].str.startswith("customop::"))]
    if POST:
        cpu_time = df_cpu["total_time"].sum()
    else:
        cpu_time = df_cpu["total_time"].sum() - icore_out_time

    # 后处理耗时
    post_time = 0

    # io算子耗时（与右边表格保持一致，减去memcpy_time）
    if 'is_io_process' in df.columns:
        df_io = df[df['is_io_process'] == True]
        io_total_time = df_io['total_time'].sum()
        io_memcpy_time = df_io['memcpy_time'].sum()
        io_time = io_total_time - io_memcpy_time  # 减去memcpy，与右边表格一致
        
        # 计算数据传入时间：主干算子的memcpy + IO算子的memcpy
        df_backbone_hardop = df[(df["op_type"] == "icraft::xir::HardOp") & (df['is_io_process'] == False)]
        backbone_memcpy = df_backbone_hardop["memcpy_time"].sum()
        data_in_time = backbone_memcpy + io_memcpy_time

        # 呈现给用户的时间
        row_names = ["数据传入时间", "icore(npu)主干时间","Icraft-CPU算子耗时","后处理耗时(自行填写)",'io算子时间']
        column_data = [data_in_time, icore_time, cpu_time, post_time, io_time]
        df_detailtime = pd.DataFrame(column_data, index=row_names,columns=['time(ms)'])
        df_detailtime['Device'] =["cdma","npu","cpu","null","npu"]
    else:
        # 呈现给用户的时间
        row_names = ["数据传入时间", "icore(npu)主干时间","数据传出时间","Icraft-CPU算子耗时","后处理耗时(自行填写)"]
        column_data = [icore_in_time, icore_time, icore_out_time, cpu_time, post_time]
        df_detailtime = pd.DataFrame(column_data, index=row_names,columns=['time(ms)'])
        df_detailtime['Device'] =[imk_list,"npu",post_list,"cpu","null"]

    # df_detailtime.columns.name = '网络各阶段耗时细则'
    df_detailtime = df_detailtime.round(4)

    # 延时和吞吐率的Excel公式（两种情况都是5行数据）
    delay_time = '=SUM(Sheet1!B4:B8)'
    fps = '=1000/MAX(Sheet1!B4:B8)'

    row_names_res = ["单帧延时(ms)","极限fps"]
    column_data_res = [delay_time, fps]
    df_res = pd.DataFrame(column_data_res, index=row_names_res,columns=['results'])
    bottleneck_formula = '=INDEX(A4:A8,MATCH(MAX(B4:B8),B4:B8,0))'
    df_res["耗时瓶颈"]=[bottleneck_formula, bottleneck_formula]
    
    # for print - 根据是否有 is_io_process 列使用不同的计算方式
    if 'is_io_process' in df.columns:
        # 有 IO 算子时：数据传入 + icore主干 + CPU + 后处理 + IO
        delay_time = data_in_time + icore_time + cpu_time + post_time + io_time
        fps = 1000/max(data_in_time, icore_time, cpu_time, post_time, io_time)
    else:
        # 没有 IO 算子时：使用完整的 5 项时间
        delay_time = icore_in_time + icore_time + icore_out_time + cpu_time + post_time
        fps = 1000/max(icore_in_time, icore_time, icore_out_time, cpu_time, post_time)
    
    df_res_print = pd.DataFrame([delay_time, fps], index=row_names_res,columns=['results'])
    df_res_print["耗时瓶颈"]=[df_detailtime['time(ms)'].idxmax(),df_detailtime['time(ms)'].idxmax()]

    # 打印细则
    def get_display_width(s):
        """计算字符串的显示宽度（中文算2个字符）"""
        width = 0
        for char in str(s):
            if '\u4e00' <= char <= '\u9fff':  # 中文字符范围
                width += 2
            else:
                width += 1
        return width
    
    def format_table(df):
        """格式化DataFrame，正确处理中文对齐"""
        lines = []
        # 表头
        headers = [''] + list(df.columns)
        col_widths = [25]  # 第一列（索引）宽度
        for col in df.columns:
            col_widths.append(max(15, get_display_width(col) + 2))
        
        # 打印表头
        header_line = ""
        for i, h in enumerate(headers):
            display_w = get_display_width(h)
            padding = col_widths[i] - display_w
            header_line += h + " " * padding
        lines.append(header_line)
        
        # 打印数据行
        for idx, row in df.iterrows():
            line = ""
            # 索引列
            idx_str = str(idx)
            display_w = get_display_width(idx_str)
            padding = col_widths[0] - display_w
            line += idx_str + " " * padding
            
            # 数据列
            for i, val in enumerate(row):
                # 格式化数值：如果是浮点数，保留4位小数
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    try:
                        val_str = f"{float(val):.4f}"
                    except:
                        val_str = str(val)
                else:
                    val_str = str(val)
                
                display_w = get_display_width(val_str)
                padding = col_widths[i+1] - display_w
                line += val_str + " " * padding
            lines.append(line)
        
        return '\n'.join(lines)
    
    print("\n统计分析结果如下:")
    print(format_table(df_detailtime))
    print("="*60)

    # 打印预估情况
    print("\n网络单帧延时、吞吐率分析结果如下：")
    print(format_table(df_res_print))
    print("="*60)

    # 写入excel
    title_row = pd.DataFrame(index=['统计分析结果如下:'])
    title_row.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow)
    df_detailtime.to_excel(writer, sheet_name=sheet_name, startcol=startcol,startrow=startrow+2)
    title_row2 = pd.DataFrame(index=['单帧延时、吞吐率分析结果如下:'])
    title_row2.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow+10)
    df_res.to_excel(writer, sheet_name=sheet_name, startcol=startcol,startrow=startrow+12)
    writer.save()

    
def analyze_time(filepath,speedMode=False):
    """
    输入:
    filepath 表示输入文件路径，通常为 modelname.xlsx(python 时间测试结果) 或 modelname.txt(C++ 时间测试结果)
    res_root="./res/"  , 表示保存文件路径
    speedMode = True or False , 表示是否开启speedmode
    输出：
    开启speedmode:
        分析结果modelname_res.xlsx
    未开启speedmode:
        分析结果modelname_res.xlsx、柱状堆叠图
    """
    # 结果路径就是输入文件所在的目录
    respath = os.path.dirname(filepath) + os.sep
    # 文件名则是输入文件的基本名称（不含扩展名）
    filename = os.path.basename(filepath).split('.')[0]
    
    filetype = detect_file_type(filepath)
    if filetype == "excel":
        input_data = pd.read_excel(filepath, sheet_name=None)
    elif filetype == "txt":
        data = []
        with open(filepath, 'r',encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                line_data = {}
                if line[:5]=="*****":
                    break
                else:
                    for item in line.split(', '):
                        key, value = item.split(': ')
                        if key == "op_type" or key == "op_name":
                            value = value.replace("Node","")
                        if value == "":
                            value  = np.nan
                        try:
                            line_data[key] = float(value)
                        except ValueError:
                            line_data[key] = value
                    data.append(line_data)
        input_data = {'Sheet1': pd.DataFrame(data)}
    else:
        print("not support current file type")
    res_name = respath + filename + "_res.xlsx"

    writer = pd.ExcelWriter(res_name, engine='openpyxl')
    max_index = 0
    if speedMode:
        for sheet_name, df_origin in input_data.items():
            df = df_origin.sort_values(by="op_id",ascending=True)
            df = df.reset_index(drop=True)
            max_index = len(df.index)

            # 统计数据传入icore时间、icore耗时、传出icore时间、纯CPU端算子耗时
            calculate_detailtime(writer,df,sheet_name,startcol=0,startrow=0)  

            # 统计各类硬算子的时间，如imk、hardop、detpost、segpost
            calculate_hardop_time(writer,df,sheet_name,startcol=4,startrow=0)

            # 统计所有op的总耗时情况
            calculate_totaltime(writer,df,sheet_name,startcol=12,startrow=0) 
    else:
        for sheet_name, df_origin in input_data.items():
            df = df_origin.sort_values(by="op_id",ascending=True)
            df = df.reset_index(drop=True)
            max_index = len(df.index)

            # 1. 统一准备主干算子数据 (df_main)
            if 'is_io_process' in df.columns:
                df_main = df[df['is_io_process'] == False].copy()
            else:
                df_main = df.copy()

            # 统计数据传入icore时间、icore耗时、传出icore时间、纯CPU端算子耗时
            calculate_detailtime(writer,df,sheet_name,startcol=0,startrow=0)

            # 统计各类硬算子的时间，如imk、hardop、detpost、segpost
            calculate_hardop_time(writer,df,sheet_name,startcol=4,startrow=0)

            # 统计所有HardOP算子细则：按照算子类别分类、统计时间
            calculate_hardop_class_time(writer,df,sheet_name,startcol=4,startrow=8)

            # 统计所有op的总耗时情况
            calculate_totaltime(writer,df,sheet_name,startcol=12,startrow=0)
            
            df_sum = df_main.groupby(["op_name"]).sum()
            op_name = [tmp.split('::')[-1] for tmp in df_sum.index]

            # 绘制各类hardop算子的耗时直方图
            draw_total_time(op_name,df_sum,filename,respath,"各类hardop算子的耗时直方图")
            
            # 绘制所有算子的耗时直方图
            df_filtered = df_main[df_main["op_type"]=="icraft::xir::HardOp"]#数据筛选,只保留op_type=hard_op
            op_id = df_filtered["op_id"]
            draw_per_op_time(op_id,df_filtered,filename,respath,"所有hardop算子的耗时直方图")

    sheet_style(writer, max_index, df)
    print("时间分析结果保存在 ", res_name)
def txt_to_excel(txt_file, excel_file):
    """
    读取指定文本文件，提取格式化的操作信息，并转换为Excel文件。
    
    Reads the specified text file, extracts formatted operation information,
    and converts it to an Excel file.
    """
    records = []
    
    try:
        with open(txt_file, 'r', encoding='utf-8') as f:
            for line in f:
                stripped_line = line.strip()
                
                # 当遇到分隔符时停止处理
                # Stop processing when the delimiter is encountered
                if stripped_line.startswith('************************************'):
                    break
                
                # 跳过空行
                # Skip empty lines
                if not stripped_line:
                    continue
                
                # 解析每一行到字典中
                # Parse each line into a dictionary
                record = {}
                parts = stripped_line.split(',')
                
                for part in parts:
                    part = part.strip()
                    if not part or ':' not in part:
                        continue
                    
                    key, value = part.split(':', 1)
                    key = key.strip()
                    value = value.strip()
                    
                    if not key:
                        continue
                    
                    # 根据键名进行类型转换
                    # Apply type conversion based on key name
                    if key == 'op_id':
                        try:
                            record[key] = int(value)
                        except (ValueError, TypeError):
                            record[key] = value
                    elif key in ['op_name', 'op_type']:
                        #--- 修改开始：去除op_type中的Node后缀(txt有Node,xlsx没有Node)
                        if key == 'op_type' and value.endswith('Node'):
                            value = value[:-4]  # 去掉末尾的"Node"（4个字符）
                        #--- 修改结束
                        record[key] = value
                    elif key in ['total_time', 'memcpy_time', 'hard_time', 'other_time']:
                        try:
                            record[key] = float(value) if value else 0.0
                        except (ValueError, TypeError):
                            record[key] = 0.0
                    elif key == 'is_io_process':
                        record[key] = value.lower() == 'true'
                    else:
                        # 处理任何其他意外键
                        # Handle any other unexpected keys
                        record[key] = value
                
                # 只添加包含数据的记录
                # Only add records that contain data
                if record:
                    records.append(record)
    
    except FileNotFoundError:
        print(f"错误：文件 '{txt_file}' 未找到。")
        return
    except Exception as e:
        print(f"读取文件 '{txt_file}' 时发生错误: {e}")
        return

    if not records:
        print(f"在 '{txt_file}' 的分隔符之前未找到有效数据。")
        return

    # 定义用户指定的列顺序
    # Define the user-specified column order
    columns = ['op_id', 'op_name', 'op_type', 'total_time', 'memcpy_time', 'hard_time', 'other_time', 'is_io_process']
    
    # 从记录列表创建DataFrame
    # Create DataFrame from the list of records
    try:
        df = pd.DataFrame(records)
        # 按指定顺序重新排列列
        # Reorder columns to match the specified order
        df = df[columns]
    except Exception as e:
        print(f"创建DataFrame时出错: {e}")
        return

    # 将DataFrame写入Excel文件
    # Write DataFrame to Excel file
    try:
        # 对.xlsx文件使用openpyxl引擎
        # Use openpyxl engine for .xlsx files
        df.to_excel(excel_file, index=False, engine='openpyxl')
        print(f"成功将 '{txt_file}' 转换为 '{excel_file}'。")
        print(f"  - 总记录数: {len(df)}")
        print(f"  - 列名: {list(df.columns)}")
    except ImportError:
        print(f"错误：写入.xlsx文件需要'openpyxl'库。请使用 'pip install openpyxl' 安装。")
    except Exception as e:
        print(f"写入Excel文件 '{excel_file}' 时出错: {e}")
